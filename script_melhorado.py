import pyautogui
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from time import sleep
import threading
import openpyxl
import os
import unicodedata

# ──────────────────────────────────────────────────────────────────
#  CONTEXTO DO SITE (Conexão Educação):
#
#  - Cada linha = um aluno
#  - Cada coluna = um dia do mês
#  - O site já vem com TODOS os dias marcados como PRESENTE (✅)
#  - Para registrar FALTA: pressionar Space (desmarca o checkbox) + Tab
#  - Para manter PRESENÇA: apenas Tab (não toca no checkbox)
#  - Dias sem aula (vazio, Cancelado, etc.): apenas Tab para pular
#  - Tab avança para o próximo DIA do mesmo aluno (esquerda → direita)
#  - Ao terminar os dias de um aluno → Enter para ir ao próximo aluno
# ──────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────
#  Paleta de Cores Moderna (Dark Mode Premium)
# ──────────────────────────────────────────────────────────────────
BG_MAIN       = "#0f172a"  # Slate 900 - Fundo Principal
BG_CARD       = "#1e293b"  # Slate 800 - Painéis/Cards
BG_INPUT      = "#0f172a"  # Slate 900 - Inputs/Listbox
PRIMARY       = "#6366f1"  # Indigo 500 - Cor Principal (Ações primárias)
PRIMARY_HOVER = "#4f46e5"  # Indigo 600
SUCCESS       = "#22c55e"  # Green 500  - Sucesso/Ativos
SUCCESS_HOVER = "#16a34a"  # Green 600
WARNING       = "#f59e0b"  # Amber 500  - Alertas/Pausar/Remanejados
WARNING_HOVER = "#d97706"  # Amber 600
DANGER        = "#ef4444"  # Red 500    - Parar/Erros/Cancelados
DANGER_HOVER  = "#dc2626"  # Red 600
TEXT_MAIN     = "#f8fafc"  # Slate 50   - Textos principais
TEXT_SEC      = "#94a3b8"  # Slate 400  - Textos secundários
TEXT_MUTED    = "#475569"  # Slate 600  - Desativados
BORDER        = "#334155"  # Slate 700  - Bordas sutis


# ──────────────────────────────────────────────────────────────────
#  Leitura do Excel (Lógica Pura - Mantida Intacta)
# ──────────────────────────────────────────────────────────────────
def ler_excel(caminho, aba, linha_inicio, linha_fim, col_inicio, col_fim, col_nome):
    """
    Lê o bloco de presenças do diário Excel e o nome dos alunos.
    Ignora alunos com status "Rem. de sala", "Cancelado" ou "Transf. de u.e.".
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[aba] if isinstance(aba, int) else wb[aba]

    alunos_ativos = []
    contadores = {
        "Ativos": 0,
        "Rem. de sala": 0,
        "Cancelado": 0,
        "Transf. de u.e.": 0
    }

    for row in range(linha_inicio, linha_fim + 1):
        nome_val = ws.cell(row=row, column=col_nome).value
        nome = str(nome_val).strip() if nome_val else f"Aluno da Linha {row}"
        
        status_aluno = "Ativos"
        nome_lower = nome.lower()
        
        if "rem. de sala" in nome_lower or "remanejado" in nome_lower:
            status_aluno = "Rem. de sala"
        elif "cancelado" in nome_lower:
            status_aluno = "Cancelado"
        elif "transf. de u.e." in nome_lower or "transferido" in nome_lower:
            status_aluno = "Transf. de u.e."

        dias = []
        for col in range(col_inicio, col_fim + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                dias.append(' ')
            else:
                s = str(val).strip().lower()
                
                if "rem. de sala" in s or "remanejado" in s:
                    status_aluno = "Rem. de sala"
                elif "cancelado" in s:
                    status_aluno = "Cancelado"
                elif "transf. de u.e." in s or "transf." in s:
                    status_aluno = "Transf. de u.e."
                
                if s == 'f':
                    dias.append('f')
                else:
                    dias.append('.')

        if status_aluno != "Ativos":
            contadores[status_aluno] += 1
        else:
            contadores["Ativos"] += 1
            alunos_ativos.append({"nome": nome, "dias": dias})
            
    def normalizar_nome(aluno):
        nome_str = aluno["nome"].lower()
        return ''.join(c for c in unicodedata.normalize('NFD', nome_str) if unicodedata.category(c) != 'Mn')
        
    alunos_ativos.sort(key=normalizar_nome)
            
    return alunos_ativos, contadores


# ──────────────────────────────────────────────────────────────────
#  Aplicação GUI (Orientada a Objetos)
# ──────────────────────────────────────────────────────────────────
class AutoFrequenciaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Frequência Automática — Conexão Educação")
        self.root.geometry("1366x950")
        self.root.minsize(850, 550)
        self.root.configure(bg=BG_MAIN)
        
        self.estado = {
            "rodando":       False,
            "pausado":       False,
            "alunos":        [],   
            "alunos_importados": [], 
            "aluno_selecionado": [], 
            "total_alunos":  0,
            "aluno_atual":   0,
            "total_celulas": 0,
            "celula_atual":  0,
        }
        
        self.setup_styles()
        self.create_ui()

    def setup_styles(self):
        style = ttk.Style(self.root)
        style.theme_use("clam")
        style.configure("Modern.Horizontal.TProgressbar",
                        troughcolor=BG_INPUT,
                        background=PRIMARY,
                        bordercolor=BG_CARD,
                        lightcolor=PRIMARY,
                        darkcolor=PRIMARY,
                        thickness=12)
        style.configure("TScrollbar",
                        troughcolor=BG_CARD,
                        background=BORDER,
                        bordercolor=BG_CARD,
                        arrowcolor=TEXT_MAIN,
                        relief="flat")

    # ──────────────────────────────────────────────────────────────────
    #  Construção da Interface
    # ──────────────────────────────────────────────────────────────────
    def create_ui(self):
        # --- HEADER ---
        header = tk.Frame(self.root, bg=BG_MAIN)
        header.pack(fill="x", padx=30, pady=(25, 15))
        
        title_frame = tk.Frame(header, bg=BG_MAIN)
        title_frame.pack(side="left")
        
        tk.Label(title_frame, text="⚡ Frequência Automática", font=("Segoe UI", 22, "bold"), 
                 bg=BG_MAIN, fg=TEXT_MAIN).pack(anchor="w")
        tk.Label(title_frame, text="Conexão Educação • Preenchimento Rápido e Seguro", 
                 font=("Segoe UI", 11), bg=BG_MAIN, fg=PRIMARY).pack(anchor="w")

        # --- MAIN GRID ---
        main_frame = tk.Frame(self.root, bg=BG_MAIN)
        main_frame.pack(fill="both", expand=True, padx=30, pady=(0, 20))
        
        main_frame.columnconfigure(0, weight=1) # Coluna Esquerda mais fina
        main_frame.columnconfigure(1, weight=2) # Coluna Direita mais larga
        main_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=0)

        # --- LEFT PANEL ---
        left_col = tk.Frame(main_frame, bg=BG_MAIN)
        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        self.create_excel_config_card(left_col)
        self.create_actions_card(left_col)

        # --- RIGHT PANEL ---
        right_col = tk.Frame(main_frame, bg=BG_MAIN)
        right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        self.create_summary_card(right_col)
        self.create_preview_card(right_col)
        
        # --- PROGRESS PANEL ---
        progress_col = tk.Frame(main_frame, bg=BG_MAIN)
        progress_col.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(5, 0))
        self.create_progress_card(progress_col)

        # --- FOOTER / STATUS BAR ---
        footer = tk.Frame(self.root, bg="#0b1120", height=40)
        footer.pack(side="bottom", fill="x")
        footer.pack_propagate(False)
        
        self.lbl_status = tk.Label(footer, text="Pronto. Configure os campos e carregue o Excel.", 
                                   bg="#0b1120", fg=TEXT_SEC, font=("Segoe UI", 10))
        self.lbl_status.pack(side="left", padx=30, pady=10)


    # ──────────────────────────────────────────────────────────────────
    #  Componentes (Cards)
    # ──────────────────────────────────────────────────────────────────
    def create_card(self, parent, title, fill="x", expand=False):
        """Cria um card com design moderno e padding consistente."""
        card = tk.Frame(parent, bg=BG_CARD)
        card.pack(fill=fill, expand=expand, pady=(0, 15))
        
        inner = tk.Frame(card, bg=BG_CARD, padx=20, pady=20)
        inner.pack(fill="both", expand=True)
        
        if title:
            lbl = tk.Label(inner, text=title, bg=BG_CARD, fg=TEXT_MAIN, font=("Segoe UI", 12, "bold"))
            lbl.pack(anchor="w", pady=(0, 15))
            
        return inner

    def create_input(self, parent, label_text, default_val, row, col):
        """Cria um label e um campo de texto moderno em um Grid."""
        lbl = tk.Label(parent, text=label_text, bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 9))
        lbl.grid(row=row, column=col, sticky="w", pady=(8, 2), padx=5)
        
        # Borda sutil de 1px
        frame_input = tk.Frame(parent, bg=BORDER, padx=1, pady=1) 
        frame_input.grid(row=row+1, column=col, sticky="we", padx=5, pady=(0, 8))
        
        entry = tk.Entry(frame_input, bg=BG_INPUT, fg=TEXT_MAIN, insertbackground=TEXT_MAIN, 
                         relief="flat", font=("Segoe UI", 10))
        entry.insert(0, default_val)
        entry.pack(fill="both", expand=True, padx=8, pady=6)
        return entry

    def create_btn(self, parent, text, bg_color, hover_color, fg_color, command):
        """Cria um botão moderno com hover states."""
        btn = tk.Button(parent, text=text, bg=bg_color, fg=fg_color, 
                        activebackground=hover_color, activeforeground=fg_color,
                        relief="flat", borderwidth=0, font=("Segoe UI", 10, "bold"), 
                        cursor="hand2", command=command, padx=16, pady=8)
        
        btn.bind("<Enter>", lambda e: btn.config(bg=hover_color) if btn['state'] != 'disabled' else None)
        btn.bind("<Leave>", lambda e: btn.config(bg=bg_color) if btn['state'] != 'disabled' else None)
        
        btn.default_bg = bg_color
        btn.hover_bg = hover_color
        return btn

    def set_btn_state(self, btn, state):
        """Atualiza o estado visual do botão (Normal / Disabled)."""
        if state == "disabled":
            btn.config(state="disabled", bg=BG_MAIN, fg=TEXT_MUTED, cursor="arrow")
        else:
            btn.config(state="normal", bg=btn.default_bg, fg=TEXT_MAIN, cursor="hand2")

    def create_stat_box(self, parent, title, val, color, col):
        """Cria uma caixa de métricas no estilo Dashboard."""
        f = tk.Frame(parent, bg=BG_INPUT, padx=10, pady=12)
        f.grid(row=0, column=col, sticky="nsew", padx=4)
        
        lbl_val = tk.Label(f, text=val, bg=BG_INPUT, fg=color, font=("Segoe UI", 18, "bold"))
        lbl_val.pack()
        lbl_title = tk.Label(f, text=title, bg=BG_INPUT, fg=TEXT_SEC, font=("Segoe UI", 9))
        lbl_title.pack()
        return lbl_val

    # ──────────────────────────────────────────────────────────────────
    #  Construção Específica dos Painéis
    # ──────────────────────────────────────────────────────────────────
    def create_excel_config_card(self, parent):
        card = self.create_card(parent, "📊 Configuração do Excel")
        
        grid_frame = tk.Frame(card, bg=BG_CARD)
        grid_frame.pack(fill="x")
        for i in range(2): grid_frame.columnconfigure(i, weight=1)
        
        self.entrada_aba      = self.create_input(grid_frame, "Aba (nome/nº)", "Abr", 0, 0)
        self.entrada_col_nome = self.create_input(grid_frame, "Coluna Nome", "3", 0, 1)
        
        self.entrada_lin_ini  = self.create_input(grid_frame, "Linha Início", "11", 2, 0)
        self.entrada_lin_fim  = self.create_input(grid_frame, "Linha Fim", "0", 2, 1)
        
        self.entrada_col_ini  = self.create_input(grid_frame, "Col. Início Dias", "4", 4, 0)
        self.entrada_col_fim  = self.create_input(grid_frame, "Col. Fim Dias", "0", 4, 1)

        file_frame = tk.Frame(card, bg=BG_CARD)
        file_frame.pack(fill="x", pady=(20, 0))
        
        btn_carregar = self.create_btn(file_frame, "📂 Carregar Excel", PRIMARY, PRIMARY_HOVER, TEXT_MAIN, self.carregar_excel)
        btn_carregar.pack(side="left")
        
        self.label_arquivo = tk.Label(file_frame, text="Nenhum arquivo", bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 9))
        self.label_arquivo.pack(side="left", padx=15)

    def create_speed_config_card(self, parent):
        card = self.create_card(parent, "⚙️ Velocidade")
        grid_frame = tk.Frame(card, bg=BG_CARD)
        grid_frame.pack(fill="x")
        grid_frame.columnconfigure(0, weight=1)
        grid_frame.columnconfigure(1, weight=1)
        
        self.entrada_delay = self.create_input(grid_frame, "Delay Dias (seg)", "0.05", 0, 0)
        self.entrada_delay_aluno = self.create_input(grid_frame, "Delay Alunos (seg)", "0.3", 0, 1)

    def create_actions_card(self, parent):
        card = self.create_card(parent, "🚀 Ações")
        btn_frame = tk.Frame(card, bg=BG_CARD)
        btn_frame.pack(fill="x", pady=(5, 0))
        
        self.btn_iniciar = self.create_btn(btn_frame, "▶ Iniciar", SUCCESS, SUCCESS_HOVER, TEXT_MAIN, self.executar_automacao)
        self.btn_iniciar.pack(side="left", fill="x", expand=True, padx=(0, 5))
        
        self.btn_pausar = self.create_btn(btn_frame, "⏸ Pausar", WARNING, WARNING_HOVER, TEXT_MAIN, self.pausar_retomar)
        self.btn_pausar.pack(side="left", fill="x", expand=True, padx=5)
        
        self.btn_parar = self.create_btn(btn_frame, "⏹ Parar", DANGER, DANGER_HOVER, TEXT_MAIN, self.parar)
        self.btn_parar.pack(side="left", fill="x", expand=True, padx=(5, 0))

        self.set_btn_state(self.btn_pausar, "disabled")
        self.set_btn_state(self.btn_parar, "disabled")

    def create_summary_card(self, parent):
        card = self.create_card(parent, "👥 Resumo do Diário")
        sum_frame = tk.Frame(card, bg=BG_CARD)
        sum_frame.pack(fill="x")
        for i in range(4): sum_frame.columnconfigure(i, weight=1)

        self.lbl_ativos = self.create_stat_box(sum_frame, "Ativos", "0", SUCCESS, 0)
        self.lbl_rem    = self.create_stat_box(sum_frame, "Remanejados", "0", WARNING, 1)
        self.lbl_canc   = self.create_stat_box(sum_frame, "Cancelados", "0", DANGER, 2)
        self.lbl_transf = self.create_stat_box(sum_frame, "Transferidos", "0", PRIMARY, 3)

    def create_preview_card(self, parent):
        card = self.create_card(parent, "👁 Alunos (Clique para (des)marcar)", expand=True, fill="both")
        
        list_container = tk.Frame(card, bg=BORDER, padx=1, pady=1)
        list_container.pack(fill="both", expand=True, pady=(5, 0))
        
        inner_list = tk.Frame(list_container, bg=BG_INPUT)
        inner_list.pack(fill="both", expand=True)

        scroll_y = ttk.Scrollbar(inner_list, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        
        self.listbox_alunos = tk.Listbox(inner_list, bg=BG_INPUT, fg=TEXT_MAIN,
                                         font=("Consolas", 11), relief="flat", borderwidth=0,
                                         selectbackground=PRIMARY, selectforeground="#ffffff",
                                         highlightthickness=0, yscrollcommand=scroll_y.set)
        self.listbox_alunos.pack(fill="both", expand=True, padx=5, pady=5)
        scroll_y.config(command=self.listbox_alunos.yview)
        
        self.listbox_alunos.bind('<Button-1>', self.toggle_selecao)

    def create_progress_card(self, parent):
        card = self.create_card(parent, "📈 Progresso")
        
        self.barra_progresso = ttk.Progressbar(card, style="Modern.Horizontal.TProgressbar",
                                               orient="horizontal", mode="determinate")
        self.barra_progresso.pack(fill="x", pady=(5, 10))
        
        self.lbl_progresso = tk.Label(card, text="Aluno 0 / 0", bg=BG_CARD, fg=TEXT_SEC, font=("Segoe UI", 10))
        self.lbl_progresso.pack()


    # ──────────────────────────────────────────────────────────────────
    #  Lógica de Interface e Eventos
    # ──────────────────────────────────────────────────────────────────
    def atualizar_status(self, msg, tipo="ok"):
        cores = {"ok": SUCCESS, "aviso": WARNING, "erro": DANGER}
        cor = cores.get(tipo, TEXT_SEC)
        self.lbl_status.config(text=msg, fg=cor)

    def carregar_excel(self):
        caminho = filedialog.askopenfilename(
            title="Selecione o diário Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
        )
        if not caminho:
            return

        try:
            aba = self.entrada_aba.get().strip()
            try: aba = int(aba)
            except ValueError: pass

            lin_ini = int(self.entrada_lin_ini.get())
            lin_fim = int(self.entrada_lin_fim.get())
            col_ini = int(self.entrada_col_ini.get())
            col_fim = int(self.entrada_col_fim.get())
            col_nome = int(self.entrada_col_nome.get())

            alunos_ativos, contadores = ler_excel(caminho, aba, lin_ini, lin_fim, col_ini, col_fim, col_nome)

            # Atualiza Dashboard
            self.lbl_ativos.config(text=str(contadores['Ativos']))
            self.lbl_rem.config(text=str(contadores['Rem. de sala']))
            self.lbl_canc.config(text=str(contadores['Cancelado']))
            self.lbl_transf.config(text=str(contadores['Transf. de u.e.']))

            # Preview na lista
            self.listbox_alunos.delete(0, tk.END)
            self.estado["alunos_importados"] = alunos_ativos
            self.estado["aluno_selecionado"] = [True] * len(alunos_ativos)
            
            for aluno in alunos_ativos:
                # ☑ (Marcado) / ☐ (Desmarcado)
                self.listbox_alunos.insert(tk.END, f" ☑  {aluno['nome']}: {''.join(aluno['dias'])}")

            nome_arquivo = os.path.basename(caminho)
            if len(nome_arquivo) > 30: nome_arquivo = nome_arquivo[:27] + "..."
            self.label_arquivo.config(text=f"📄 {nome_arquivo}")
            
            self.atualizar_status(f"✅ Excel carregado. {len(alunos_ativos)} alunos ativos encontrados.", "ok")

        except Exception as e:
            messagebox.showerror("Erro ao ler Excel", str(e))
            self.atualizar_status("Erro ao carregar o arquivo.", "erro")

    def toggle_selecao(self, event):
        if not self.estado.get("alunos_importados"):
            return
            
        idx = self.listbox_alunos.nearest(event.y)
        if idx >= 0:
            bbox = self.listbox_alunos.bbox(idx)
            if bbox:
                y_start, y_end = bbox[1], bbox[1] + bbox[3]
                if y_start <= event.y <= y_end:
                    self.estado["aluno_selecionado"][idx] = not self.estado["aluno_selecionado"][idx]
                    aluno = self.estado["alunos_importados"][idx]
                    
                    marca = "☑" if self.estado["aluno_selecionado"][idx] else "☐"
                    novo_texto = f" {marca}  {aluno['nome']}: {''.join(aluno['dias'])}"
                    
                    self.listbox_alunos.delete(idx)
                    self.listbox_alunos.insert(idx, novo_texto)
                    
                    if not self.estado["aluno_selecionado"][idx]:
                        self.listbox_alunos.itemconfig(idx, foreground=TEXT_MUTED)
                    else:
                        self.listbox_alunos.itemconfig(idx, foreground=TEXT_MAIN)


    # ──────────────────────────────────────────────────────────────────
    #  Automação
    # ──────────────────────────────────────────────────────────────────
    def executar_automacao(self):
        if "alunos_importados" in self.estado and self.estado["alunos_importados"]:
            self.estado["alunos"] = [
                aluno for i, aluno in enumerate(self.estado["alunos_importados"])
                if self.estado["aluno_selecionado"][i]
            ]

        if not self.estado["alunos"]:
            messagebox.showwarning("Atenção", "Nenhum aluno válido selecionado para processamento.")
            return

        self.estado["total_alunos"]  = len(self.estado["alunos"])
        self.estado["aluno_atual"]   = 0
        self.estado["total_celulas"] = sum(len(a["dias"]) for a in self.estado["alunos"])
        self.estado["celula_atual"]  = 0
        self.estado["rodando"]       = True
        self.estado["pausado"]       = False

        self.barra_progresso["maximum"] = self.estado["total_alunos"]
        self.barra_progresso["value"]   = 0
        self.lbl_progresso.config(text=f"Iniciando: 0 / {self.estado['total_alunos']}")

        self.set_btn_state(self.btn_iniciar, "disabled")
        self.set_btn_state(self.btn_pausar, "normal")
        self.btn_pausar.config(text="⏸ Pausar")
        self.set_btn_state(self.btn_parar, "normal")

        thread = threading.Thread(target=self._loop_automacao, daemon=True)
        thread.start()

    def _loop_automacao(self):
        self.atualizar_status("⏳ Aguardando 4 segundos — clique na janela do site...", "aviso")
        sleep(4)
        pyautogui.press('tab')
        
        for idx_aluno, aluno in enumerate(self.estado["alunos"]):
            if not self.estado["rodando"]:
                break

            while self.estado["pausado"] and self.estado["rodando"]:
                sleep(0.2)

            self.root.after(0, lambda i=idx_aluno, n=aluno["nome"]: self._atualizar_status_aluno(i, n))

            # ── dias do aluno ──
            for ch in aluno["dias"]:
                if not self.estado["rodando"]:
                    break
                while self.estado["pausado"] and self.estado["rodando"]:
                    sleep(0.2)

                if ch == 'f':
                    pyautogui.press('space')
                    pyautogui.press('tab')
                else:
                    pyautogui.press('tab')

                self.estado["celula_atual"] += 1

            # ── próximo aluno ──
            if self.estado["rodando"]:
                self.estado["aluno_atual"] = idx_aluno + 1
                self.root.after(0, self._atualizar_barra)

        self.estado["rodando"] = False
        self.root.after(0, self._finalizar_ui)

    def _atualizar_status_aluno(self, idx, nome_aluno):
        total   = self.estado["total_alunos"]
        cel_at  = self.estado["celula_atual"]
        cel_tot = self.estado["total_celulas"]
        pct     = int(cel_at / cel_tot * 100) if cel_tot else 0
        
        texto_prog = f"{nome_aluno} ({idx + 1} / {total})"
        if len(texto_prog) > 50:
            texto_prog = texto_prog[:47] + "..."
            
        self.lbl_progresso.config(text=texto_prog)
        self.atualizar_status(f"Processando: {nome_aluno} ({pct}% concluído)", "ok")

    def _atualizar_barra(self):
        self.barra_progresso["value"] = self.estado["aluno_atual"]

    def _finalizar_ui(self):
        self.set_btn_state(self.btn_iniciar, "normal")
        self.set_btn_state(self.btn_pausar, "disabled")
        self.btn_pausar.config(text="⏸ Pausar")
        self.set_btn_state(self.btn_parar, "disabled")
        
        total = self.estado["total_alunos"]
        atual = self.estado["aluno_atual"]
        if atual == total and total > 0:
            self.atualizar_status(f"✅ Concluído! {total} alunos processados com sucesso.", "ok")
        else:
            self.atualizar_status(f"⛔ Interrompido. {atual} de {total} alunos processados.", "aviso")

    def pausar_retomar(self):
        if not self.estado["rodando"]:
            return
        
        self.estado["pausado"] = not self.estado["pausado"]
        if self.estado["pausado"]:
            self.btn_pausar.config(text="▶ Retomar")
            self.atualizar_status("⏸ Automação Pausada.", "aviso")
        else:
            self.btn_pausar.config(text="⏸ Pausar")
            self.atualizar_status("▶ Retomando automação...", "ok")

    def parar(self):
        self.estado["rodando"] = False
        self.estado["pausado"] = False
        self.atualizar_status("⏹ Parando automação...", "aviso")


# ──────────────────────────────────────────────────────────────────
#  Execução
# ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app = AutoFrequenciaApp(root)
    root.mainloop()
