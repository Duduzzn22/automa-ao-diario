import pyautogui
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from time import sleep
import threading
import openpyxl
import os
import unicodedata

# ──────────────────────────────────────────────────────────────────
#  CONTEXTO DO SITE (Conexão Educação):
#
#  - Cada linha = um aluno
#  - Cada coluna = um dia do mês
#  - O site já vem com TODOS os dias marcados como PRESENTE (✅)
#  - Para registrar FALTA: pressionar Space (desmarca o checkbox) + Tab
#  - Para manter PRESENÇA: apenas Tab (não toca no checkbox)
#  - Dias sem aula (vazio, Cancelado, etc.): apenas Tab para pular
#  - Tab avança para o próximo DIA do mesmo aluno (esquerda → direita)
#  - Ao terminar os dias de um aluno → Enter para ir ao próximo aluno
#
#  MELHORIAS:
#  - Alunos exibidos pelo NOME real (como está no Excel)
#  - Ordenação ALFABÉTICA A→Z (igual ao site Conexão Educação)
#  - Alunos com "Cancelado", "Rem. de Sala" ou similares são IGNORADOS
#    e não contabilizados no progresso
# ──────────────────────────────────────────────────────────────────

# Termos que indicam que o aluno deve ser completamente ignorado
TERMOS_IGNORAR_ALUNO = [
    "cancelado", "rem. de Sala", "transf. de u.e.",
    ]
estado = {
    "rodando":       False,
    "pausado":       False,
    "alunos":        [],   # lista de dicts: {nome, dias}
    "total_alunos":  0,
    "aluno_atual":   0,
    "total_celulas": 0,
    "celula_atual":  0,
}


# ──────────────────────────────────────────────────────────────────
#  Utilitários
# ──────────────────────────────────────────────────────────────────
def remover_acentos(texto):
    """Remove acentos para ordenação alfabética correta (Á → A, etc.)."""
    return unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode('ascii')


def aluno_deve_ser_ignorado(nome):
    """Verifica se o aluno possui algum marcador de exclusão no nome."""
    nome_lower = nome.lower()
    return any(termo in nome_lower for termo in TERMOS_IGNORAR_ALUNO)


def celula_deve_ser_ignorada(valor_str):
    """Verifica se uma célula do Excel indica situação especial de ignorar."""
    v = valor_str.strip().lower()
    return any(termo in v for termo in TERMOS_IGNORAR_ALUNO)


# ──────────────────────────────────────────────────────────────────
#  Leitura do Excel
# ──────────────────────────────────────────────────────────────────
def ler_excel(caminho, aba, linha_nome, linha_inicio, linha_fim,
              col_nome, col_inicio, col_fim):
    """
    Lê o bloco de presenças do diário Excel.

    Parâmetros (todos 1-based):
        caminho      : caminho do arquivo .xlsx
        aba          : nome da aba (ex: 'Mar') ou índice numérico (ex: 0)
        linha_nome   : coluna onde estão os NOMES dos alunos (ex: 3 = col C)
        linha_inicio : primeira linha de aluno (ex: 11)
        linha_fim    : última linha de aluno
        col_nome     : coluna com o nome do aluno (ex: 3 = coluna C)
        col_inicio   : primeira coluna de dias (ex: 4 = coluna D)
        col_fim      : última coluna de dias

    Retorna lista de dicts ordenada ALFABETICAMENTE:
        [
          {'nome': 'ÁLVARO ALMEIDA CARNEIRO', 'dias': ['.', 'f', ' ', ...]},
          {'nome': 'ANA VITÓRIA ...',          'dias': [...]},
          ...
        ]
        Alunos com "Cancelado" / "Rem. de Sala" no nome são excluídos.

    Legenda dos dias:
        '.' = presente
        'f' = falta
        ' ' = pular (vazio, Cancelado, qualquer outro texto na célula do dia)
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[aba] if isinstance(aba, int) else wb[aba]

    alunos_brutos = []

    for row in range(linha_inicio, linha_fim + 1):
        # ── Nome do aluno ──
        val_nome = ws.cell(row=row, column=col_nome).value
        nome = str(val_nome).strip() if val_nome is not None else ""

        # Pula linhas sem nome
        if not nome or nome == "None":
            continue

        # Ignora aluno cancelado / removido (pelo nome)
        if aluno_deve_ser_ignorado(nome):
            continue

        v = 0
        # ── Dias de frequência ──
        dias = []
        for col in range(col_inicio, col_fim + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                dias.append(' ')
            else:
                s = str(val).strip()
                s_lower = s.lower()
                if s_lower == '':
                    dias.append(' ')
                elif s_lower == 'f':
                    dias.append('f')
                elif s == '.' or s_lower == "a":
                    dias.append('.')
                elif celula_deve_ser_ignorada(s):
                    v = 1
                    dias.append(' ')
                else:
                    dias.append(' ')
        if v == 0:
            alunos_brutos.append({'nome': nome, 'dias': dias})

    # ── Ordenação alfabética A → Z (ignorando acentos) ──
    alunos_ordenados = sorted(
        alunos_brutos,
        key=lambda a: remover_acentos(a['nome'].upper())
    )

    return alunos_ordenados


def carregar_excel():
    caminho = filedialog.askopenfilename(
        title="Selecione o diário Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")]
    )
    if not caminho:
        return

    try:
        aba = entrada_aba.get().strip()
        try:
            aba = int(aba)
        except ValueError:
            pass

        col_nome  = int(entrada_col_nome.get())
        lin_ini   = int(entrada_lin_ini.get())
        lin_fim   = int(entrada_lin_fim.get())
        col_ini   = int(entrada_col_ini.get())
        col_fim   = int(entrada_col_fim.get())

        alunos = ler_excel(
            caminho, aba,
            col_nome, lin_ini, lin_fim,
            col_nome, col_ini, col_fim
        )

        # ── Preview na área de texto ──
        area_texto.config(state="normal")
        area_texto.delete("1.0", tk.END)
        for i, aluno in enumerate(alunos, start=1):
            dias_str = ''.join(aluno['dias'])
            faltas   = dias_str.count('f')
            area_texto.insert(
                tk.END,
                f"{i:02d}. {aluno['nome']:<45}  |  {dias_str}  "
                f"({'F:'+str(faltas) if faltas else 'sem faltas'})\n"
            )
        area_texto.config(state="disabled")

        total_cel = sum(len(a['dias']) for a in alunos)
        label_arquivo.config(text=f"📄 {os.path.basename(caminho)}")
        atualizar_status(
            f"✅ {len(alunos)} alunos (ordenados A→Z) — {total_cel} células carregadas.",
            "ok"
        )
        estado["alunos"] = alunos

    except Exception as e:
        messagebox.showerror("Erro ao ler Excel", str(e))


# ──────────────────────────────────────────────────────────────────
#  Automação
# ──────────────────────────────────────────────────────────────────
def executar_automacao():
    if not estado["alunos"]:
        messagebox.showwarning(
            "Atenção", "Nenhum dado carregado.\nCarregue o Excel primeiro."
        )
        return

    estado["total_alunos"]  = len(estado["alunos"])
    estado["aluno_atual"]   = 0
    estado["total_celulas"] = sum(len(a['dias']) for a in estado["alunos"])
    estado["celula_atual"]  = 0
    estado["rodando"]       = True
    estado["pausado"]       = False

    barra_progresso["maximum"] = estado["total_alunos"]
    barra_progresso["value"]   = 0
    label_progresso.config(text=f"Aluno 0 / {estado['total_alunos']}")

    btn_iniciar.config(state="disabled")
    btn_pausar.config(state="normal", text="⏸  Pausar")
    btn_parar.config(state="normal")

    thread = threading.Thread(target=_loop_automacao, daemon=True)
    thread.start()


def _loop_automacao():
    delay       = float(entrada_delay.get())
    delay_aluno = float(entrada_delay_aluno.get())

    atualizar_status("Aguardando 4 s — clique na janela do site...", "aviso")
    sleep(4)

    for idx_aluno, aluno in enumerate(estado["alunos"]):

        if not estado["rodando"]:
            break

        while estado["pausado"] and estado["rodando"]:
            sleep(0.2)

        nome_aluno = aluno['nome']
        janela.after(0, lambda i=idx_aluno, n=nome_aluno: _atualizar_status_aluno(i, n))

        # ── dias do aluno ──
        for ch in aluno['dias']:
            if not estado["rodando"]:
                break
            while estado["pausado"] and estado["rodando"]:
                sleep(0.2)

            if ch == 'f':
                pyautogui.press('space')   # desmarca o checkbox de presente
                sleep(delay)
                pyautogui.press('tab')
            else:
                pyautogui.press('tab')     # presente ou vazio: só avança

            estado["celula_atual"] += 1
            sleep(delay)

        # ── próximo aluno ──
        if estado["rodando"]:
            estado["aluno_atual"] = idx_aluno + 1
            janela.after(0, _atualizar_barra)
            sleep(delay_aluno)

    estado["rodando"] = False
    janela.after(0, _finalizar_ui)


def _atualizar_status_aluno(idx, nome):
    total   = estado["total_alunos"]
    cel_at  = estado["celula_atual"]
    cel_tot = estado["total_celulas"]
    pct     = int(cel_at / cel_tot * 100) if cel_tot else 0
    label_progresso.config(text=f"Aluno {idx + 1} / {total}  —  {nome}")
    atualizar_status(
        f"👤 [{idx + 1}/{total}] {nome}  ({pct}% concluído)", "ok"
    )


def _atualizar_barra():
    barra_progresso["value"] = estado["aluno_atual"]


def _finalizar_ui():
    btn_iniciar.config(state="normal")
    btn_pausar.config(state="disabled", text="⏸  Pausar")
    btn_parar.config(state="disabled")
    total = estado["total_alunos"]
    atual = estado["aluno_atual"]
    if atual == total:
        atualizar_status(f"✅ Concluído! {total} alunos processados.", "ok")
    else:
        atualizar_status(f"⛔ Interrompido no aluno {atual + 1} de {total}.", "aviso")


def pausar_retomar():
    if not estado["rodando"]:
        return
    estado["pausado"] = not estado["pausado"]
    if estado["pausado"]:
        btn_pausar.config(text="▶  Retomar")
        atualizar_status("⏸ Pausado.", "aviso")
    else:
        btn_pausar.config(text="⏸  Pausar")
        atualizar_status("▶ Retomando…", "ok")


def parar():
    estado["rodando"] = False
    estado["pausado"] = False


def atualizar_status(msg, tipo="ok"):
    cores = {"ok": "#2ecc71", "aviso": "#f39c12", "erro": "#e74c3c"}
    label_status.config(text=msg, foreground=cores.get(tipo, "#ecf0f1"))


# ──────────────────────────────────────────────────────────────────
#  Interface gráfica
# ──────────────────────────────────────────────────────────────────
janela = tk.Tk()
janela.title("Frequência Automática — Conexão Educação")
janela.resizable(False, False)
janela.configure(bg="#1a1a2e")

style = ttk.Style(janela)
style.theme_use("clam")
style.configure("TFrame",       background="#1a1a2e")
style.configure("TLabel",       background="#1a1a2e", foreground="#ecf0f1", font=("Segoe UI", 10))
style.configure("TEntry",       fieldbackground="#16213e", foreground="#ecf0f1", insertcolor="#ecf0f1")
style.configure("TLabelframe",  background="#1a1a2e", foreground="#a29bfe", font=("Segoe UI", 10, "bold"))
style.configure("TLabelframe.Label", background="#1a1a2e", foreground="#a29bfe")
style.configure("cyan.Horizontal.TProgressbar",
                troughcolor="#16213e", background="#00cec9", thickness=18)

BG    = "#1a1a2e"
PANEL = "#16213e"
ROXO  = "#a29bfe"
VERDE = "#00cec9"
LRNJ  = "#fd9644"
AZUL  = "#0f3460"

# ── Título ──
tk.Label(janela, text="📋  Frequência Automática — Conexão Educação",
         bg=BG, fg=ROXO, font=("Segoe UI", 14, "bold")).pack(padx=20, pady=(16, 6), anchor="w")

tk.Label(janela,
         text="✦ Alunos ordenados A→Z  ✦  Cancelado / Rem. de Sala ignorados automaticamente",
         bg=BG, fg="#636e72", font=("Segoe UI", 9)).pack(padx=20, pady=(0, 8), anchor="w")

# ── Configuração Excel ──
frame_excel = ttk.LabelFrame(janela, text=" 📊  Configuração do Excel ", padding=12)
frame_excel.pack(fill="x", padx=20, pady=6)

cfg = ttk.Frame(frame_excel)
cfg.pack(fill="x")

# linha 1
ttk.Label(cfg, text="Aba (nome ou 0,1,2…):").grid(row=0, column=0, sticky="w", padx=(0,4), pady=3)
entrada_aba = ttk.Entry(cfg, width=10); entrada_aba.insert(0, "Mar")
entrada_aba.grid(row=0, column=1, sticky="w", padx=(0,24), pady=3)

ttk.Label(cfg, text="Coluna dos NOMES (núm.):").grid(row=0, column=2, sticky="w", padx=(0,4), pady=3)
entrada_col_nome = ttk.Entry(cfg, width=6); entrada_col_nome.insert(0, "3")
entrada_col_nome.grid(row=0, column=3, sticky="w", pady=3)

# linha 2
ttk.Label(cfg, text="Linha início dos alunos:").grid(row=1, column=0, sticky="w", padx=(0,4), pady=3)
entrada_lin_ini = ttk.Entry(cfg, width=6); entrada_lin_ini.insert(0, "11")
entrada_lin_ini.grid(row=1, column=1, sticky="w", padx=(0,24), pady=3)

ttk.Label(cfg, text="Linha fim dos alunos:").grid(row=1, column=2, sticky="w", padx=(0,4), pady=3)
entrada_lin_fim = ttk.Entry(cfg, width=6); entrada_lin_fim.insert(0, "45")
entrada_lin_fim.grid(row=1, column=3, sticky="w", pady=3)

# linha 3
ttk.Label(cfg, text="Coluna início dos dias:").grid(row=2, column=0, sticky="w", padx=(0,4), pady=3)
entrada_col_ini = ttk.Entry(cfg, width=6); entrada_col_ini.insert(0, "4")
entrada_col_ini.grid(row=2, column=1, sticky="w", padx=(0,24), pady=3)

ttk.Label(cfg, text="Coluna fim dos dias:").grid(row=2, column=2, sticky="w", padx=(0,4), pady=3)
entrada_col_fim = ttk.Entry(cfg, width=6); entrada_col_fim.insert(0, "34")
entrada_col_fim.grid(row=2, column=3, sticky="w", pady=3)

# botão carregar
fb = ttk.Frame(frame_excel); fb.pack(fill="x", pady=(10, 0))
tk.Button(fb, text="📂  Carregar Excel", command=carregar_excel,
          bg=AZUL, fg="white", font=("Segoe UI", 10, "bold"),
          relief="flat", padx=12, pady=5, cursor="hand2").pack(side="left")
label_arquivo = ttk.Label(fb, text="nenhum arquivo selecionado", foreground="#636e72")
label_arquivo.pack(side="left", padx=12)

# ── Preview ──
frame_prev = ttk.LabelFrame(
    janela,
    text=" 👁  Preview  ( . = presente   f = falta   espaço = pular ) — ordenado A→Z ",
    padding=10
)
frame_prev.pack(fill="x", padx=20, pady=6)

scroll_y = tk.Scrollbar(frame_prev)
scroll_y.pack(side="right", fill="y")
scroll_x = tk.Scrollbar(frame_prev, orient="horizontal")
scroll_x.pack(side="bottom", fill="x")

area_texto = tk.Text(
    frame_prev, height=9, bg=PANEL, fg="#ecf0f1",
    font=("Consolas", 9), relief="flat", wrap="none",
    yscrollcommand=scroll_y.set,
    xscrollcommand=scroll_x.set,
    state="disabled"
)
area_texto.pack(fill="x")
scroll_y.config(command=area_texto.yview)
scroll_x.config(command=area_texto.xview)

# ── Velocidade ──
frame_vel = ttk.LabelFrame(janela, text=" ⚙️  Velocidade ", padding=12)
frame_vel.pack(fill="x", padx=20, pady=6)

fv = ttk.Frame(frame_vel); fv.pack(fill="x")
ttk.Label(fv, text="Delay entre dias (seg):").grid(row=0, column=0, sticky="w", padx=(0,6))
entrada_delay = ttk.Entry(fv, width=7); entrada_delay.insert(0, "0.05")
entrada_delay.grid(row=0, column=1, sticky="w", padx=(0,30))

ttk.Label(fv, text="Delay entre alunos (seg):").grid(row=0, column=2, sticky="w", padx=(0,6))
entrada_delay_aluno = ttk.Entry(fv, width=7); entrada_delay_aluno.insert(0, "0.3")
entrada_delay_aluno.grid(row=0, column=3, sticky="w")

# Botões
fb2 = ttk.Frame(frame_vel); fb2.pack(fill="x", pady=(12, 0))
btn_iniciar = tk.Button(fb2, text="▶  Iniciar", command=executar_automacao,
                        bg=VERDE, fg="#1a1a2e", font=("Segoe UI", 11, "bold"),
                        relief="flat", padx=16, pady=7, cursor="hand2")
btn_iniciar.pack(side="left", padx=(0, 8))

btn_pausar = tk.Button(fb2, text="⏸  Pausar", command=pausar_retomar, state="disabled",
                       bg=LRNJ, fg="#1a1a2e", font=("Segoe UI", 11, "bold"),
                       relief="flat", padx=16, pady=7, cursor="hand2")
btn_pausar.pack(side="left", padx=(0, 8))

btn_parar = tk.Button(fb2, text="⏹  Parar", command=parar, state="disabled",
                      bg="#d63031", fg="white", font=("Segoe UI", 11, "bold"),
                      relief="flat", padx=16, pady=7, cursor="hand2")
btn_parar.pack(side="left")

# ── Progresso ──
frame_prog = ttk.LabelFrame(janela, text=" 📈  Progresso ", padding=10)
frame_prog.pack(fill="x", padx=20, pady=6)
barra_progresso = ttk.Progressbar(frame_prog, style="cyan.Horizontal.TProgressbar",
                                  orient="horizontal", mode="determinate")
barra_progresso.pack(fill="x")
label_progresso = ttk.Label(frame_prog, text="Aluno 0 / 0",
                             font=("Segoe UI", 9))
label_progresso.pack(pady=(4, 0))

# ── Status ──
label_status = ttk.Label(janela,
                         text="Pronto. Configure os campos e carregue o Excel.",
                         foreground=VERDE, font=("Segoe UI", 10))
label_status.pack(pady=(6, 18))

janela.mainloop()
